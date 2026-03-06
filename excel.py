from selenium import webdriver
from selenium.webdriver.ie.service import Service
from webdriver_manager.core.manager import DriverManager
import openpyxl

Dict={}
book=openpyxl.load_workbook('/Users/arpitamajumdar/Downloads/spreadsheet.xlsx')
sheet=book.active
print(sheet)
print(sheet.max_row)
print(sheet.max_column)
print(sheet['A4'].value)
cell=sheet.cell(row=2,column=2)
print(cell.value)
update=sheet.cell(row=2,column=2).value="Arpita"
print(update)
print("*********************")
for i in range(1,sheet.max_row+1): # row prining
    if sheet.cell(row=i,column=1).value=="c":
        for j in range(2,sheet.max_column+1): # column printing
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
print(Dict)

for i in range(1,sheet.max_row+1):
    for j in range(2,sheet.max_column+1):
        print(sheet.cell(row=i,column=j).value)



