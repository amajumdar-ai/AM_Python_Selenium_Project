import time

import webdriver_manager
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl



# function to update the column value in downloaded excel sheet

def excel_file_upload(file_path,col_name,search_term,new_value):

   excel_file=openpyxl.load_workbook(file_path)
   sheet=excel_file.active
   row=sheet.max_row
   dict_item = {}
   column=sheet.max_column
   for i in range(1, column+1):
     if sheet.cell(row=1,column=i).value==col_name:
       dict_item["col"]=i
   for i in range(1,row+1):
     for j in range(2,column+1):
        if sheet.cell(row=i,column=j).value==search_term:
            dict_item["row"]=i
   sheet.cell(row=dict_item["row"], column=dict_item["col"]).value = new_value
   excel_file.save(file_path)


service=Service(ChromeDriverManager().install())
driver=webdriver.Chrome(service=service)

fruit_name="Apple"
new_value="900"
search_term="price"
file_path="/Users/arpitamajumdar/Downloads/download.xlsx" #downloaded file path
driver.maximize_window()
driver.get("https://rahulshettyacademy.com/upload-download-test/")
time.sleep(5)

# download file
driver.find_element(By.ID,"downloadButton").click()

# UPDATE THE COLUMN VALUE IN EXCEL
excel_file_upload(file_path, "price", fruit_name, new_value)

# upload file
driver.find_element(By.ID,"fileinput").send_keys(file_path)


# success toaster test
wait = WebDriverWait(driver,5)
toast_locator = (By.CSS_SELECTOR,".Toastify__toast-body div:nth-child(2)")
wait.until(expected_conditions.visibility_of_element_located(toast_locator))
print(driver.find_element(*toast_locator).text)
time.sleep(5)

# price value column retrival
price_column=driver.find_element(By.XPATH,"//div[text()='Price']").get_attribute("data-column-id")
actual_price=driver.find_element(By.XPATH,"//div[text()='"+fruit_name+"']/parent::div/parent::div/div[@id='cell-"+price_column+"-undefined']").text
print(actual_price)

# assertion to check the actual price and updated new price on the excel sheet
assert actual_price==new_value
print("assertion passed")
driver.find_element(By.ID,"fileinput").send_keys(file_path)
time.sleep(5)